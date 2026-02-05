from pathlib import Path
from typing import Generator, Callable
from openpyxl import load_workbook


class ExcelParser:
    def __init__(
        self,
        file_path: str | Path,
        sheet_name: str | None = None,
        progress_callback: Callable[[int], None] | None = None,
    ):
        self.file_path = Path(file_path)
        self.sheet_name = sheet_name
        self.progress_callback = progress_callback
        self._row_count: int | None = None
        self._wb = None
        self._ws = None

    def _get_row_count(self, ws) -> int:
        return ws.max_row - 1

    def _read_rows(
        self,
        ws,
        start_row: int = 0,
    ) -> Generator[dict[str, str], None, None]:
        if ws.max_row <= 1:
            return
        headers = [cell.value for cell in ws[1]]
        for row_index in range(2, ws.max_row + 1):
            if row_index - 1 < start_row:
                continue
            row_data = {}
            for col_index, header in enumerate(headers, start=1):
                cell_value = ws.cell(row=row_index, column=col_index).value
                row_data[header] = str(cell_value) if cell_value is not None else ""
            yield row_data
            if self.progress_callback:
                self.progress_callback(row_index - 1)

    def parse(
        self,
        start_row: int = 0,
    ) -> tuple[Generator[dict[str, str], None, None], int]:
        self._wb = load_workbook(self.file_path, read_only=True, data_only=True)
        self._ws = self._wb[self.sheet_name] if self.sheet_name else self._wb.active
        if start_row == 0:
            self._row_count = self._get_row_count(self._ws)

        def generator():
            try:
                yield from self._read_rows(self._ws, start_row)
            finally:
                if self._wb:
                    self._wb.close()
                    self._wb = None
                    self._ws = None

        return generator(), self._row_count or 0

    def get_row_count(self) -> int:
        if self._row_count is None:
            wb = load_workbook(self.file_path, read_only=True, data_only=True)
            ws = wb.active
            self._row_count = self._get_row_count(ws)
            wb.close()
        return self._row_count

    def get_sheets(self) -> list[str]:
        wb = load_workbook(self.file_path, read_only=True)
        sheets = wb.sheetnames
        wb.close()
        return sheets
