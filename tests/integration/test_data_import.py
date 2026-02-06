import pytest
import tempfile
import os
from datetime import datetime, timezone
from unittest.mock import AsyncMock, MagicMock, patch


from src.domains.data_import.enums import ImportStatus
from src.domains.data_import.models import DataImportRecord
from src.domains.data_import.service import ImportService
from src.domains.data_import.mapping import FieldMapper
from src.domains.data_import.validator import ValidationService


class TestDataImportIntegration:
    @pytest.fixture
    def mock_redis(self):
        redis = AsyncMock()
        redis.set = AsyncMock()
        redis.get = AsyncMock(return_value=None)
        redis.exists = AsyncMock(return_value=False)
        return redis

    @pytest.fixture
    def temp_csv_file(self):
        with tempfile.NamedTemporaryFile(
            mode="w", suffix=".csv", delete=False, newline=""
        ) as f:
            f.write("order_no,amount,order_date,quantity\n")
            f.write("O001,100.50,2024-01-15,2\n")
            f.write("O002,200.00,2024-01-16,3\n")
            f.write("O003,50.25,2024-01-17,1\n")
            temp_path = f.name
        yield temp_path
        if os.path.exists(temp_path):
            os.unlink(temp_path)

    @pytest.fixture
    def temp_excel_file(self, temp_csv_file):
        import csv
        import openpyxl

        rows = []
        with open(temp_csv_file, "r", encoding="utf-8") as f:
            reader = csv.DictReader(f)
            rows = list(reader)

        wb = openpyxl.Workbook()
        ws = wb.active
        if rows:
            headers = list(rows[0].keys())
            for col_idx, header in enumerate(headers, 1):
                ws.cell(row=1, column=col_idx, value=header)
            for row_idx, row in enumerate(rows, 2):
                for col_idx, header in enumerate(headers, 1):
                    value = row.get(header)
                    if value and header in ["amount", "quantity"]:
                        try:
                            value = float(value)
                        except (ValueError, TypeError):
                            pass
                    ws.cell(row=row_idx, column=col_idx, value=value)

        temp_xlsx = temp_csv_file.replace(".csv", ".xlsx")
        wb.save(temp_xlsx)
        yield temp_xlsx
        if os.path.exists(temp_xlsx):
            os.unlink(temp_xlsx)


class TestFullImportWorkflow:
    @pytest.fixture
    def mock_session(self):
        session = AsyncMock()
        session.commit = AsyncMock()
        session.flush = AsyncMock()
        return session

    @pytest.fixture
    def mock_redis(self):
        redis = AsyncMock()
        redis.set = AsyncMock()
        redis.get = AsyncMock(return_value=None)
        redis.exists = AsyncMock(return_value=False)
        return redis

    @pytest.mark.asyncio
    async def test_complete_import_workflow(self, mock_session, mock_redis):
        service = ImportService(session=mock_session, redis=mock_redis)

        with patch.object(
            service.repo, "create", new_callable=AsyncMock
        ) as mock_create:
            mock_record = MagicMock(spec=DataImportRecord)
            mock_record.id = 1
            mock_record.file_name = "orders.csv"
            mock_record.file_size = 1024
            mock_record.status = ImportStatus.PENDING
            mock_record.created_at = datetime.now(timezone.utc)
            mock_record.field_mapping = None
            mock_create.return_value = mock_record

            _ = await service.upload_file(
                file_path="/uploads/orders.csv",
                file_name="orders.csv",
                file_size=1024,
                file_type="csv",
                data_source_id=1,
                batch_no="IMP-12345678",
                user_id=1,
            )

            mock_create.assert_called_once()
            call_args = mock_create.call_args[0][0]
            assert call_args["file_name"] == "orders.csv"
            assert call_args["batch_no"] == "IMP-12345678"

    @pytest.mark.asyncio
    async def test_field_mapping_workflow(self):
        mapper = FieldMapper()
        mapper.set_target_fields(["order_id", "amount", "order_date", "quantity"])

        source_fields = ["order_no", "total_amount", "下单日期", "num"]
        _ = mapper.auto_map(source_fields, required_fields=["order_id"])

        mapping_dict = mapper.get_mapping_dict()
        assert "order_no" in mapping_dict

        row = {
            "order_no": "O123",
            "total_amount": "100.50",
            "下单日期": "2024-01-15",
            "num": "5",
        }

        transformed = mapper.transform_data(row)

        assert "order_id" in transformed
        assert "amount" in transformed

    @pytest.mark.asyncio
    async def test_validation_workflow(self):
        rows = [
            {
                "order_id": "O001",
                "amount": 100.50,
                "order_date": "2024-01-15",
                "quantity": 2,
            },
            {
                "order_id": "O002",
                "amount": 200.00,
                "order_date": "2024-01-16",
                "quantity": 3,
            },
            {"order_id": "", "amount": -50.00, "order_date": "invalid", "quantity": 1},
        ]

        results = ValidationService.validate("order", rows)
        summary = ValidationService.validate_and_summarize("order", rows)

        assert "total_rows" in summary
        assert "passed" in summary
        assert "failed" in summary
        assert "errors_by_field" in summary
        assert summary["total_rows"] == 3
        assert summary["passed"] >= 1
        assert summary["failed"] >= 1

        failed_result = next((r for r in results if r.status.value == "fail"), None)
        assert failed_result is not None
        assert len(failed_result.errors) > 0


class TestMappingToValidationPipeline:
    def test_mapping_then_validation_pipeline(self):
        mapper = FieldMapper()
        mapper.set_target_fields(["order_id", "amount", "order_date"])
        mapper.add_manual_mapping("ord_no", "order_id")
        mapper.add_manual_mapping("total_amount", "amount")
        mapper.add_manual_mapping("order_dt", "order_date")

        source_rows = [
            {"ord_no": "O001", "total_amount": "100.50", "order_dt": "2024-01-15"},
            {"ord_no": "O002", "total_amount": "200.00", "order_dt": "2024-01-16"},
            {"ord_no": "O003", "total_amount": "50.25", "order_dt": "2024-01-17"},
        ]

        mapped_rows = [mapper.transform_data(row) for row in source_rows]

        assert len(mapped_rows) == 3
        assert mapped_rows[0]["order_id"] == "O001"
        assert mapped_rows[0]["order_date"] == "2024-01-15"

        _ = ValidationService.validate("order", mapped_rows)
        summary = ValidationService.validate_and_summarize("order", mapped_rows)

        assert summary["total_rows"] == 3
        assert summary["passed"] == 3


class TestDataImportModels:
    def test_import_record_model_creation(self):
        from src.domains.data_import.enums import FileType

        record = DataImportRecord(
            id=1,
            batch_no="IMP-TEST123",
            file_name="test.csv",
            file_type=FileType.CSV,
            file_size=2048,
            file_path="/uploads/test.csv",
            status=ImportStatus.PENDING,
            total_rows=0,
            success_rows=0,
            failed_rows=0,
            created_by_id=1,
        )

        assert record.id == 1
        assert record.batch_no == "IMP-TEST123"
        assert record.status == ImportStatus.PENDING
        assert record.file_type == FileType.CSV

    def test_import_status_enum(self):
        assert ImportStatus.PENDING.value == "pending"
        assert ImportStatus.PROCESSING.value == "processing"
        assert ImportStatus.SUCCESS.value == "success"
        assert ImportStatus.FAILED.value == "failed"
        assert ImportStatus.PARTIAL.value == "partial"


class TestErrorHandling:
    @pytest.fixture
    def mock_session(self):
        session = AsyncMock()
        session.commit = AsyncMock()
        return session

    @pytest.fixture
    def mock_redis(self):
        redis = AsyncMock()
        redis.set = AsyncMock()
        return redis

    @pytest.mark.asyncio
    async def test_parse_nonexistent_record(self, mock_session, mock_redis):
        service = ImportService(session=mock_session, redis=mock_redis)

        with patch.object(
            service.repo, "get_by_id", new_callable=AsyncMock, return_value=None
        ):
            with pytest.raises(ValueError, match="Import record not found"):
                await service.parse_file(999)

    @pytest.mark.asyncio
    async def test_apply_mapping_nonexistent_record(self, mock_session, mock_redis):
        service = ImportService(session=mock_session, redis=mock_redis)

        with patch.object(
            service.repo, "get_by_id", new_callable=AsyncMock, return_value=None
        ):
            with pytest.raises(ValueError, match="Import record not found"):
                await service.apply_mapping(
                    import_id=999, mappings={"old": "new"}, target_fields=["new"]
                )

    @pytest.mark.asyncio
    async def test_confirm_import_failed_record(self, mock_session, mock_redis):
        service = ImportService(session=mock_session, redis=mock_redis)

        mock_record = MagicMock()
        mock_record.status = ImportStatus.FAILED
        mock_record.field_mapping = None

        with patch.object(
            service.repo, "get_by_id", new_callable=AsyncMock, return_value=mock_record
        ):
            with pytest.raises(ValueError, match="Import has failed"):
                await service.confirm_import(1, [{"field": "value"}])
