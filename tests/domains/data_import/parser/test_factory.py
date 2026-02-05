import pytest
import tempfile
import os
import gc
from openpyxl import Workbook


def test_factory_creates_csv_parser():
    from src.domains.data_import.parser.factory import FileParser

    with tempfile.NamedTemporaryFile(mode="w", suffix=".csv", delete=False) as f:
        f.write("id,name\n1,test\n")
        temp_path = f.name

    try:
        parser = FileParser(temp_path)
        assert hasattr(parser._parser, "_detect_encoding")
    finally:
        os.unlink(temp_path)


def test_factory_creates_excel_parser():
    from src.domains.data_import.parser.factory import FileParser

    temp_path = tempfile.mktemp(suffix=".xlsx")
    wb = Workbook()
    ws = wb.active
    assert ws is not None
    ws["A1"] = "id"
    wb.save(temp_path)
    del wb
    gc.collect()

    try:
        parser = FileParser(temp_path)
        assert hasattr(parser._parser, "get_sheets")
    finally:
        os.unlink(temp_path)


def test_validate_file_rejects_unsupported():
    from src.domains.data_import.parser.factory import FileParser

    is_valid, msg = FileParser.validate_file("/path/to/file.txt")
    assert is_valid is False
    assert "Unsupported file format" in msg


def test_validate_file_accepts_csv():
    from src.domains.data_import.parser.factory import FileParser

    with tempfile.NamedTemporaryFile(mode="w", suffix=".csv", delete=False) as f:
        f.write("id,name\n1,test\n")
        temp_path = f.name

    try:
        is_valid, msg = FileParser.validate_file(temp_path)
        assert is_valid is True
        assert msg == ""
    finally:
        os.unlink(temp_path)


def test_validate_file_accepts_xlsx():
    from src.domains.data_import.parser.factory import FileParser

    temp_path = tempfile.mktemp(suffix=".xlsx")
    wb = Workbook()
    ws = wb.active
    assert ws is not None
    ws["A1"] = "id"
    wb.save(temp_path)
    del wb
    gc.collect()

    try:
        is_valid, msg = FileParser.validate_file(temp_path)
        assert is_valid is True
        assert msg == ""
    finally:
        os.unlink(temp_path)


def test_validate_file_rejects_nonexistent():
    from src.domains.data_import.parser.factory import FileParser

    is_valid, msg = FileParser.validate_file("/nonexistent/path/file.csv")
    assert is_valid is False
    assert "File not found" in msg


def test_factory_parse_csv():
    from src.domains.data_import.parser.factory import FileParser

    with tempfile.NamedTemporaryFile(mode="w", suffix=".csv", delete=False) as f:
        f.write("id,name\n1,Alice\n2,Bob\n")
        temp_path = f.name

    try:
        parser = FileParser(temp_path)
        rows = list(parser.parse())

        assert len(rows) == 2
        assert rows[0] == {"id": "1", "name": "Alice"}
    finally:
        os.unlink(temp_path)


def test_factory_parse_excel():
    from src.domains.data_import.parser.factory import FileParser

    temp_path = tempfile.mktemp(suffix=".xlsx")
    wb = Workbook()
    ws = wb.active
    assert ws is not None
    ws["A1"] = "id"
    ws["B1"] = "name"
    ws["A2"] = "1"
    ws["B2"] = "Alice"
    ws["A3"] = "2"
    ws["B3"] = "Bob"
    wb.save(temp_path)
    del wb
    gc.collect()

    try:
        parser = FileParser(temp_path)
        rows = list(parser.parse())

        assert len(rows) == 2
        assert rows[0] == {"id": "1", "name": "Alice"}
    finally:
        os.unlink(temp_path)


def test_factory_get_row_count_csv():
    from src.domains.data_import.parser.factory import FileParser

    with tempfile.NamedTemporaryFile(mode="w", suffix=".csv", delete=False) as f:
        f.write("id,name\n1,Alice\n2,Bob\n3,Charlie\n")
        temp_path = f.name

    try:
        parser = FileParser(temp_path)
        count = parser.get_row_count()

        assert count == 3
    finally:
        os.unlink(temp_path)


def test_factory_get_row_count_excel():
    from src.domains.data_import.parser.factory import FileParser

    temp_path = tempfile.mktemp(suffix=".xlsx")
    wb = Workbook()
    ws = wb.active
    assert ws is not None
    ws["A1"] = "id"
    ws["A2"] = "1"
    ws["A3"] = "2"
    ws["A4"] = "3"
    wb.save(temp_path)
    del wb
    gc.collect()

    try:
        parser = FileParser(temp_path)
        count = parser.get_row_count()

        assert count == 3
    finally:
        os.unlink(temp_path)


def test_factory_get_sheets_csv():
    from src.domains.data_import.parser.factory import FileParser

    with tempfile.NamedTemporaryFile(mode="w", suffix=".csv", delete=False) as f:
        f.write("id,name\n1,test\n")
        temp_path = f.name

    try:
        parser = FileParser(temp_path)
        sheets = parser.get_sheets()

        assert sheets is None
    finally:
        os.unlink(temp_path)


def test_factory_get_sheets_excel():
    from src.domains.data_import.parser.factory import FileParser

    temp_path = tempfile.mktemp(suffix=".xlsx")
    wb = Workbook()
    wb.create_sheet("Data1")
    wb.create_sheet("Data2")
    wb.save(temp_path)
    del wb
    gc.collect()

    try:
        parser = FileParser(temp_path)
        sheets = parser.get_sheets()

        assert sheets is not None
        assert "Data1" in sheets
        assert "Data2" in sheets
    finally:
        os.unlink(temp_path)


def test_factory_supported_formats():
    from src.domains.data_import.parser.factory import FileParser

    assert ".csv" in FileParser.SUPPORTED_FORMATS
    assert ".xlsx" in FileParser.SUPPORTED_FORMATS
    assert ".txt" not in FileParser.SUPPORTED_FORMATS


def test_factory_raises_for_unknown_format():
    from src.domains.data_import.parser.factory import FileParser

    with pytest.raises(ValueError, match="Unsupported file format"):
        FileParser("/path/to/file.unknown")


def test_factory_with_progress_callback():
    from src.domains.data_import.parser.factory import FileParser

    with tempfile.NamedTemporaryFile(mode="w", suffix=".csv", delete=False) as f:
        f.write("id,name\n1,Alice\n2,Bob\n")
        temp_path = f.name

    try:
        progress = []
        parser = FileParser(
            temp_path, progress_callback=lambda row: progress.append(row)
        )
        list(parser.parse())

        assert len(progress) == 2
        assert 1 in progress
        assert 2 in progress
    finally:
        os.unlink(temp_path)


def test_factory_with_sheet_name():
    from src.domains.data_import.parser.factory import FileParser

    temp_path = tempfile.mktemp(suffix=".xlsx")
    wb = Workbook()
    ws1 = wb.create_sheet("MainData")
    ws2 = wb.create_sheet("Other")
    ws1["A1"] = "id"
    ws1["A2"] = "1"
    ws2["A1"] = "x"
    ws2["A2"] = "y"
    wb.save(temp_path)
    del wb
    gc.collect()

    try:
        parser = FileParser(temp_path, sheet_name="MainData")
        rows = list(parser.parse())

        assert len(rows) == 1
        assert rows[0]["id"] == "1"
    finally:
        os.unlink(temp_path)
