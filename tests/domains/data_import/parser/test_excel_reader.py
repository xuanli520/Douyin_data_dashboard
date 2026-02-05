import tempfile
import os
import gc
from openpyxl import Workbook


def _create_excel_file(data: dict):
    temp_path = tempfile.mktemp(suffix=".xlsx")
    wb = Workbook()
    ws = wb.active
    if ws is None:
        ws = wb.create_sheet()
    for row_idx, row_data in enumerate(data, start=1):
        for col_idx, (key, value) in enumerate(row_data.items(), start=1):
            cell = ws.cell(row=row_idx, column=col_idx)
            cell.value = value
    wb.save(temp_path)
    del wb
    gc.collect()
    return temp_path


def test_excel_parser_get_sheets():
    from src.domains.data_import.parser.excel_reader import ExcelParser

    temp_path = tempfile.mktemp(suffix=".xlsx")
    wb = Workbook()
    wb.create_sheet("Sheet2")
    sheet1 = wb.create_sheet("Sheet1")
    sheet1["A1"] = "id"
    sheet1["B1"] = "name"
    wb.save(temp_path)
    del wb
    gc.collect()

    try:
        parser = ExcelParser(temp_path)
        sheets = parser.get_sheets()
        assert "Sheet1" in sheets
        assert "Sheet2" in sheets
    finally:
        os.unlink(temp_path)


def test_excel_parser_reads_rows():
    from src.domains.data_import.parser.excel_reader import ExcelParser

    temp_path = tempfile.mktemp(suffix=".xlsx")
    wb = Workbook()
    ws = wb.active
    if ws is None:
        ws = wb.create_sheet()
    ws["A1"] = "id"
    ws["B1"] = "name"
    ws["A2"] = "1"
    ws["B2"] = "Alice"
    ws["A3"] = "2"
    ws["B3"] = "Bob"
    wb.save(temp_path)
    del wb
    gc.collect()

    try:
        parser = ExcelParser(temp_path)
        generator, count = parser.parse()
        rows = list(generator)

        assert count == 2
        assert len(rows) == 2
        assert rows[0] == {"id": "1", "name": "Alice"}
        assert rows[1] == {"id": "2", "name": "Bob"}
    finally:
        os.unlink(temp_path)


def test_excel_parser_start_row():
    from src.domains.data_import.parser.excel_reader import ExcelParser

    temp_path = tempfile.mktemp(suffix=".xlsx")
    wb = Workbook()
    ws = wb.active
    if ws is None:
        ws = wb.create_sheet()
    ws["A1"] = "id"
    ws["B1"] = "name"
    ws["A2"] = "1"
    ws["B2"] = "Alice"
    ws["A3"] = "2"
    ws["B3"] = "Bob"
    wb.save(temp_path)
    del wb
    gc.collect()

    try:
        parser = ExcelParser(temp_path)
        generator, count = parser.parse(start_row=2)
        rows = list(generator)

        assert len(rows) == 1
        assert rows[0] == {"id": "2", "name": "Bob"}
    finally:
        os.unlink(temp_path)


def test_excel_parser_get_row_count():
    from src.domains.data_import.parser.excel_reader import ExcelParser

    temp_path = tempfile.mktemp(suffix=".xlsx")
    wb = Workbook()
    ws = wb.active
    if ws is None:
        ws = wb.create_sheet()
    ws["A1"] = "id"
    ws["A2"] = "1"
    ws["A3"] = "2"
    ws["A4"] = "3"
    wb.save(temp_path)
    del wb
    gc.collect()

    try:
        parser = ExcelParser(temp_path)
        count = parser.get_row_count()

        assert count == 3
    finally:
        os.unlink(temp_path)


def test_excel_parser_with_sheet_name():
    from src.domains.data_import.parser.excel_reader import ExcelParser

    temp_path = tempfile.mktemp(suffix=".xlsx")
    wb = Workbook()
    ws1 = wb.create_sheet("MainData")
    ws2 = wb.create_sheet("Other")
    ws1["A1"] = "id"
    ws1["A2"] = "1"
    ws2["A1"] = "x"
    ws2["A2"] = "y"
    wb.save(temp_path)
    del wb
    gc.collect()

    try:
        parser = ExcelParser(temp_path, sheet_name="MainData")
        generator, count = parser.parse()
        rows = list(generator)

        assert count == 1
        assert rows[0]["id"] == "1"
    finally:
        os.unlink(temp_path)


def test_excel_parser_handles_none_values():
    from src.domains.data_import.parser.excel_reader import ExcelParser

    temp_path = tempfile.mktemp(suffix=".xlsx")
    wb = Workbook()
    ws = wb.active
    if ws is None:
        ws = wb.create_sheet()
    ws["A1"] = "id"
    ws["B1"] = "value"
    ws["A2"] = "1"
    ws["B2"] = None
    ws["A3"] = "2"
    ws["B3"] = "data"
    wb.save(temp_path)
    del wb
    gc.collect()

    try:
        parser = ExcelParser(temp_path)
        generator, count = parser.parse()
        rows = list(generator)

        assert rows[0]["value"] == ""
        assert rows[1]["value"] == "data"
    finally:
        os.unlink(temp_path)


def test_excel_parser_get_sheets_returns_all():
    from src.domains.data_import.parser.excel_reader import ExcelParser

    temp_path = tempfile.mktemp(suffix=".xlsx")
    wb = Workbook()
    wb.create_sheet("First")
    wb.create_sheet("Second")
    wb.create_sheet("Third")
    wb.save(temp_path)
    del wb
    gc.collect()

    try:
        parser = ExcelParser(temp_path)
        sheets = parser.get_sheets()

        assert len(sheets) == 4
        assert "First" in sheets
        assert "Second" in sheets
        assert "Third" in sheets
    finally:
        os.unlink(temp_path)
